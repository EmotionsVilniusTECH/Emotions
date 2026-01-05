# -*- coding: utf-8 -*-

import pandas as pd
import statsmodels.api as sm
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.drawing.image import Image as OpenpyxlImage
import os
import matplotlib.pyplot as plt
import matplotlib.image as mpimg
import numpy as np
from fpdf import FPDF, YPos, XPos
import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.base import MIMEBase
from email import encoders
from io import BytesIO
import warnings
# Ignoruojame perspejimus, susijusius su "PyFPDF" ir "fpdf2" konfliktu
warnings.filterwarnings("ignore", category=UserWarning)
 
 
def braizyti_radara(pavadinimai, vidurkiai):
    plt.rcParams['font.family'] = 'sans-serif'
    plt.rcParams['font.sans-serif'] = ['DejaVu Sans', 'Arial', 'Tahoma'] 
    plt.rcParams['axes.unicode_minus'] = False
    N = len(pavadinimai)
    
    vidurkiai_uzdari = vidurkiai + vidurkiai[:1]
    
    angles = [n / float(N) * 2 * np.pi for n in range(N)]
    angles += angles[:1]
 
    fig = plt.figure(figsize=(4, 4)) 
    
    try:
        failo_kelias = os.path.join('AUXIL', 'Fonas1.png')
        fonas = mpimg.imread(failo_kelias)
        ax_bg = fig.add_axes([0, 0, 1, 1], zorder=-2)
        ax_bg.imshow(fonas, aspect='auto')
        ax_bg.axis('off')
    except FileNotFoundError:
        print("Ispejimas: Fono paveikslelis 'Fonas1.png' nerastas. Braizoma be fono.")
 
    ax = fig.add_subplot(111, polar=True)
 
    colors = plt.colormaps.get_cmap('rainbow')
    
    for i in range(N):
        ax.fill(
            [angles[i], angles[i+1], 0, 0],
            [vidurkiai_uzdari[i], vidurkiai_uzdari[i+1], 0, 0],
            color=colors(i / (N - 1)),
            alpha=0.8,
            edgecolor='none'
        )
    
    ax.plot(angles, vidurkiai_uzdari, linewidth=2, linestyle='solid', label='Vidurkiai', marker='o', color='black')
 
    ax.set_xticks(angles[:-1])
    ax.set_xticklabels([f'{pav}\n({vid:.4f})' for pav, vid in zip(pavadinimai, vidurkiai)], fontsize=10)
 
    max_val = max(vidurkiai)
    ax.set_ylim(0, max_val * 1.1)
    ax.set_rgrids(np.arange(0, max_val * 1.1, 0.1), fontsize=9, angle=90)
    
    ax.grid(True, color='gray', linewidth=0.5, linestyle='--')
    ax.spines['polar'].set_visible(False)
    
    return fig
 
def calculate_elasticity(df, dependent_var, independent_vars):
    elasticities = {}
    y_mean = df[dependent_var].mean()
    
    for var in independent_vars:
        try:
            X = sm.add_constant(df[var])
            y = df[dependent_var]
            model = sm.OLS(y, X).fit()
            b = model.params[var]
            x_mean = df[var].mean()
            elasticity = b * (x_mean / y_mean)
            elasticities[var] = elasticity
        except Exception:
            elasticities[var] = None
    
    return elasticities
 
def create_and_save_time_plot(df_with_time, dependent_var, save_path):
    if 'time' not in df_with_time.columns or dependent_var not in df_with_time.columns:
        return None
    
    plot_filename = os.path.join(os.path.dirname(save_path), f'{dependent_var}_time_plot.png')
    
    try:
        plt.style.use('seaborn-v0_8-whitegrid')
        fig, ax = plt.subplots(figsize=(12, 7))
 
        df_with_time[dependent_var] = pd.to_numeric(df_with_time[dependent_var], errors='coerce')
        df_with_time_clean = df_with_time.dropna(subset=[dependent_var])
 
        if df_with_time_clean.empty:
            return None
 
        x_data = df_with_time_clean.index
        y_data = df_with_time_clean[dependent_var]
        ax.plot(x_data, y_data, label=dependent_var, linestyle='-')
        
        x_ticks_locs = ax.get_xticks()
        x_tick_labels = []
        
        for tick in x_ticks_locs:
            if 0 <= int(tick) < len(df_with_time):
                real_time = df_with_time.iloc[int(tick)]['time']
                if pd.notna(real_time):
                    time_only = str(real_time).split(' ')[-1]
                    x_tick_labels.append(time_only)
                else:
                    x_tick_labels.append('')
            else:
                x_tick_labels.append('')
 
        ax.set_xticks(x_ticks_locs)
        ax.set_xticklabels(x_tick_labels, rotation=0, ha='center')
        
        if not df_with_time.empty:
            ax.set_xlim(df_with_time.index.min(), df_with_time.index.max())
 
        ax.set_xlabel('Time (hh:mm:ss)')
        ax.set_ylabel(dependent_var)
        ax.set_title(f'Dependence of {dependent_var} values on time')
        ax.legend()
        
        plt.tight_layout()
        plt.savefig(plot_filename)
        plt.close()
        
        return plot_filename
        
    except Exception as e:
        return None
 
def create_and_save_pdf(results, excel_path, df_full):
    pdf = FPDF(orientation='P', unit='mm', format='A4')
    pdf_filename = os.path.splitext(excel_path)[0] + '.pdf'
    temp_plot_paths = []
 
    try:
        pdf.add_font("Arial", "", "C:/Windows/Fonts/Arial.ttf")
        pdf.add_font("Arial", "B", "C:/Windows/Fonts/Arialbd.ttf")
        
        for dependent_var, res in results.items():
            pdf.add_page()
            
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(0, 10, f'Regression Results for: {dependent_var}', new_x=XPos.LMARGIN, new_y=YPos.NEXT, align='C')
            pdf.ln(5)
 
            pdf.set_font("Arial", '', 10)
            pdf.cell(50, 6, f'R2 = {res["r2"]:.4f}', new_x=XPos.RIGHT)
            pdf.cell(0, 6, f'p = {res["f_pvalue"]:.4f}', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
            pdf.ln(2)
 
            # Create table in PDF
            pdf.set_font("Arial", 'B', 8)
            col_widths = [35, 30, 20, 30, 30]
            cell_height = 6
 
            pdf.cell(col_widths[0], cell_height, 'Variables', 1, new_x=XPos.RIGHT)
            pdf.cell(col_widths[1], cell_height, 'Coeficient B', 1, new_x=XPos.RIGHT)
            pdf.cell(col_widths[2], cell_height, 'p', 1, new_x=XPos.RIGHT)
            pdf.cell(col_widths[3], cell_height, 'Std. beta', 1, new_x=XPos.RIGHT)
            pdf.cell(col_widths[4], cell_height, 'Elasticity', 1, new_x=XPos.LMARGIN, new_y=YPos.NEXT)
 
            pdf.set_font("Arial", '', 8)
            variables = res['model_params'].index.tolist()
            b_values = res['model_params'].values.tolist()
            p_values = res['p_values'].values.tolist()
 
            std_beta_dict = res['beta_standardized'].to_dict()
            elasticity_dict = res['elasticities']
 
            for i, var_name in enumerate(variables):
                b_val = b_values[i]
                p_val = p_values[i]
                std_beta_val = std_beta_dict.get(var_name, None)
                elasticity_val = elasticity_dict.get(var_name, None)
 
                std_beta_str = f'{std_beta_val:.4f}' if pd.notna(std_beta_val) else ''
                elasticity_str = f'{elasticity_val:.4f}' if pd.notna(elasticity_val) else ''
 
                # Left alignment for variable names
                pdf.cell(col_widths[0], cell_height, str(var_name), 1, align='L', new_x=XPos.RIGHT)
                # Right alignment for numeric values
                pdf.cell(col_widths[1], cell_height, f'{b_val:.4f}', 1, align='R', new_x=XPos.RIGHT)
                pdf.cell(col_widths[2], cell_height, f'{p_val:.4f}', 1, align='R', new_x=XPos.RIGHT)
                pdf.cell(col_widths[3], cell_height, std_beta_str, 1, align='R', new_x=XPos.RIGHT)
                pdf.cell(col_widths[4], cell_height, elasticity_str, 1, align='R', new_x=XPos.LMARGIN, new_y=YPos.NEXT)
 
            if 'time' in df_full.columns:
                plot_path = create_and_save_time_plot(df_full, dependent_var, excel_path)
                if plot_path:
                    temp_plot_paths.append(plot_path)
                    pdf.ln(10)
                    pdf.image(plot_path, x=20, y=pdf.get_y(), w=150)
        
        pdf.output(pdf_filename)
    
    except Exception as e:
        pass
    finally:
        for path in temp_plot_paths:
            try:
                os.remove(path)
            except OSError:
                pass

# 1. Konfigūracija: Teksto Vertimo Žodynai
# Visi teksto vertimai el. laiskui, sugrupuoti pagal kalbos kodą (lang_sel)
TRANSLATIONS = {
    "IT": {
        "THANK_YOU": "Grazie per aver partecipato alle nostre attività!",
        "HOW_TO_READ_GRAPHS": "Come leggere questi grafici? I grafici rappresentano quanto hai provato le differenti emozioni, in media, durante il video. Più il punto si avvicina al bordo, più hai provato l’emozione corrispondente.",
        "FACIAL_EMOTIONS": "Emozioni rilevate tramite riconoscimento facciale",
        "VOICE_EMOTIONS": "Emozioni rilevate tramite riconoscimento vocale",
        "ENVIRONMENTAL_PARAMS": "Parametri ambientali",
        "STAY_UPDATED": "Resta aggiornato/a sul progetto europeo AFFECTS!",
        "FOLLOW_US": "Seguici sui social media e rimani al corrente sulla nostra ricerca, sui prossimi eventi e attività! ",
    },
    "EN": {
        "THANK_YOU": "Thank you for participating in our activities!",
        "HOW_TO_READ_GRAPHS": "How do you read these graphs? The graphs represent how much you experienced different emotions, on average, during the video. The closer the point is to the edge, the more you experienced the corresponding emotion.",
        "FACIAL_EMOTIONS": "Emotions detected through facial recognition",
        "VOICE_EMOTIONS": "Emotions detected through voice recognition",
        "ENVIRONMENTAL_PARAMS": "Environmental parameters",
        "STAY_UPDATED": "Stay up to date on the European AFFECTS project!",
        "FOLLOW_US": "Follow us on social media and stay up-to-date on our research, upcoming events, and activities!",
    },
    "LT": {
        "THANK_YOU": "Dėkojame, kad dalyvaujate mūsų veiklose!",
        "HOW_TO_READ_GRAPHS": "Kaip skaityti šiuos grafikus? Grafikai rodo, kiek vidutiniškai patyrėte skirtingų emocijų žiūrėdami vaizdo įrašą. Kuo arčiau taško krašto, tuo labiau patyrėte atitinkamą emociją.",
        "FACIAL_EMOTIONS": "Emocijos, aptiktos veido išraiškų analizės pagalba",
        "VOICE_EMOTIONS": "Emocijos, aptiktos balso analizės pagalba",
        "ENVIRONMENTAL_PARAMS": "Aplinkos parametrai",
        "STAY_UPDATED": "Sekite naujausią informaciją apie Europos AFFECTS projektą!",
        "FOLLOW_US": "Sekite mus socialiniuose tinkluose ir gaukite naujausią informaciją apie mūsų tyrimus, artėjančius renginius ir veiklas!",
    },
    "GE": {
        "THANK_YOU": "Vielen Dank für Ihre Teilnahme an unseren Aktivitäten!",
        "HOW_TO_READ_GRAPHS": "Wie liest man diese Diagramme? Die Diagramme zeigen, wie stark Sie im Durchschnitt verschiedene Emotionen während des Videos empfunden haben. Je näher der Punkt am Rand liegt, desto stärker haben Sie die entsprechende Emotion erlebt.",
        "FACIAL_EMOTIONS": "Emotionen, die durch Gesichtserkennung erkannt werden",
        "VOICE_EMOTIONS": "Emotionen, die durch Spracherkennung erkannt werden",
        "ENVIRONMENTAL_PARAMS": "Umweltparameter",
        "STAY_UPDATED": "Bleiben Sie über das europäische AFFECTS-Projekt auf dem Laufenden!",
        "FOLLOW_US": "Folgen Sie uns in den sozialen Medien und bleiben Sie über unsere Forschung, bevorstehende Veranstaltungen und Aktivitäten auf dem Laufenden!",
    },
    "FR": {
        "THANK_YOU": "Merci d'avoir participé à nos activités !",
        "HOW_TO_READ_GRAPHS": "Comment interpréter ces graphiques ? Ils représentent l’intensité moyenne des différentes émotions ressenties pendant le visionnage de la vidéo. Plus le point est proche du bord, plus l’émotion correspondante est intense.",
        "FACIAL_EMOTIONS": "Émotions détectées par reconnaissance faciale",
        "VOICE_EMOTIONS": "Émotions détectées par reconnaissance vocale",
        "ENVIRONMENTAL_PARAMS": "Paramètres environnementaux",
        "STAY_UPDATED": "Restez informé(e) sur le projet européen AFFECTS !",
        "FOLLOW_US": "Suivez-nous sur les réseaux sociaux et restez au courant de nos recherches, de nos événements à venir et de nos activités !",
    },
}

# 2. HTML Šablonas
HTML_TEMPLATE = """
<html>
  <body>
    <p><b>{THANK_YOU}</b></></p>
    <br>
    <p style="font-size:16px;">{confidence}</p>
    <table style="width:100%;">
        <tr>
            <td style="text-align: center;" colspan="2">{HOW_TO_READ_GRAPHS}<br></td>                
        </tr>
        <tr>
            <td style="text-align: center;"><img src="cid:grafikas_id_0" alt="Grafikas 1"><p>{FACIAL_EMOTIONS}</p></td>
            <td style="text-align: center;"><img src="cid:grafikas_id_1" alt="Grafikas 2"><p>{FACIAL_EMOTIONS}</p></td>
        </tr>
        <tr>
            <td style="text-align: center;"><img src="cid:grafikas_id_2" alt="Grafikas 3"><p>{VOICE_EMOTIONS}</p></td>
            <td style="text-align: center;"><img src="cid:grafikas_id_3" alt="Grafikas 4"><p>{VOICE_EMOTIONS}</p></td>
        </tr>
        <tr>
            <td colspan="2" style="text-align: center;"><img src="cid:grafikas_id_4" alt="Grafikas 5" style="display: block; margin: 0 auto;"><p style="text-align: center;">{ENVIRONMENTAL_PARAMS}</p></td>
        </tr>
    </table>
    <p style="margin-bottom: 0;">{STAY_UPDATED}</p>
  </body>
</html>
"""
 
def send_email_with_attachments(to_email, subject, body, attachment_paths, radar_image_data_list, confidence, lang_sel):
    """Siuncia el. laiska su priedais ir grafiku per Gmail."""
    
    password = ""
    from_email = ""
    
   
    if not password:
        print("Klaida: Nerastas slaptazodis. Patikrinkite, ar irasete ji i koda.")
        return False
        
    print("Bandome prisijungti prie Gmail SMTP serverio...")
 
    # Sukuriamas pagrindinis konteineris
    msg = MIMEMultipart('related')
    msg['From'] = from_email
    msg['To'] = to_email
    msg['Subject'] = subject
 
    # Sukuriamas 'alternative' konteineris, skirtas paprastam tekstui ir HTML
    alternative_part = MIMEMultipart('alternative')
    
    # NAUJAS BLOKAS: HTML kūrimas naudojant žodyną ir šabloną
    # Pasirenkama vertimų bazė pagal kalbą. Jei kalbos nėra, naudojama anglų (EN).
    selected_translation = TRANSLATIONS.get(lang_sel, TRANSLATIONS["EN"])
    
    # Sukuriamas HTML kontekstas, įtraukiant dinaminę "confidence" reikšmę
    html_context = selected_translation.copy()
    html_context['confidence'] = confidence
    
    # Sukuriamas galutinis HTML turinys naudojant šabloną ir kontekstą
    html_body = HTML_TEMPLATE.format(**html_context)

    # Pridedame paprastą tekstą ir HTML i alternative_part
    alternative_part.attach(MIMEText(body, 'plain', 'utf-8'))
    alternative_part.attach(MIMEText(html_body, 'html', 'utf-8'))

    # Pridedame alternative_part prie related konteinerio
    msg.attach(alternative_part)
 
    # Pridedame grafiku paveikslelius kaip susijusias dalis
    for i, image_data in enumerate(radar_image_data_list):
        cid = f'grafikas_id_{i}'
        image = MIMEImage(image_data.read())
        image.add_header('Content-ID', f'<{cid}>')
        image.add_header('Content-Disposition', 'inline')
        msg.attach(image)
 
    # Pridedame kitus priedus
    for path in attachment_paths:
        if not os.path.exists(path):
            print(f"Klaida: Failas nerastas - {path}")
            continue
 
        part = MIMEBase('application', 'octet-stream')
        with open(path, 'rb') as file:
            part.set_payload(file.read())
        
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', f"attachment; filename= {os.path.basename(path)}")
        msg.attach(part)
 
    try:
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(from_email, password)
        server.send_message(msg)
        print(f"El. laiskas sekmingai issiustas gavejui {to_email}.")
        server.quit()
        return True
 
    except smtplib.SMTPAuthenticationError as e:
        print(f"Originalus klaidos pranesimas: {e}")
        return False
 
    except Exception as e:
        print(f"Ivyko bendra klaida siunciant el. laiska: {e}")
        return False
 
def normalizuoti_duomenis(df, stulpeliai):
    # Sukuriame df_temp tik su tais stulpeliais, kuriuos turime pradiniame df
    df_temp = df.copy()

    for stulpelis in stulpeliai:
        # Pirmiausia patikriname, ar stulpelis egzistuoja df_temp
        if stulpelis not in df_temp.columns:
            df_temp[stulpelis] = 0

        min_val = df_temp[stulpelis].min()
        max_val = df_temp[stulpelis].max()

        if max_val - min_val > 0:
            df_temp[stulpelis] = (df_temp[stulpelis] - min_val) / (max_val - min_val)
        else:
            df_temp[stulpelis] = 0.5

    return df_temp[stulpeliai] 
def main():
    """Pagrindine programos funkcija, kuri gaus failo kelia is komandines eilutes."""
    
    if len(sys.argv) < 4:
        sys.exit(1)
 
    csv_file_path = sys.argv[1]
    excel_to_save_path = sys.argv[2]
    lang_sel=sys.argv[3]
    
    try:
        df_full = pd.read_csv(csv_file_path, delimiter=',', decimal='.')
        df_full.rename(columns={'sad': 'sad_x'}, inplace=True)
        
        # 1. Isaugojame pradini duomenu remeli i Excel faila, nurodyta antruoju argumentu
        with pd.ExcelWriter(excel_to_save_path, engine='openpyxl') as writer:
            df_full.to_excel(writer, sheet_name='Duomenys', index=False)
            
        df = df_full.copy()
        
    except Exception as e:
        print(f"Klaida nuskaitant faila '{csv_file_path}': {e}")
        sys.exit(1)
 
    # Sukuriamas pirmas grafikas
    if lang_sel == 'IT':
        spa_list1 = ['Attivazione', 'Neutro', 'Rabbia', 'Tristezza', 'Disgusto', 'Paura']
    if lang_sel == 'EN':
        spa_list1 = ['Arousal', 'Neutral', 'Angry', 'Sad', 'Disgusted', 'Scared']
    if lang_sel == 'LT':
        spa_list1 = ['Susijaudinimas', 'Neutralumas', 'Pyktis', 'Liūdesys', 'Pasibjaurėjimas', 'Išgąstis']
    if lang_sel == 'GE':
        spa_list1 = ['Erregung', 'Neutral', 'Wütend', 'Traurig', 'Ekelhaft', 'Ängstlich']
    if lang_sel == 'FR':
        spa_list1 = ['Éveil', 'Neutre', 'Colère', 'Triste', 'Dégoût', 'Effrayé']
    pav_list1 = ['arousal', 'neutral', 'angry', 'sad_x', 'disgusted', 'scared']
    vid_list1 = df[pav_list1].mean().tolist()
    radar_fig1 = braizyti_radara(spa_list1, vid_list1)
    
    # Sukuriamas antras grafikas
    if lang_sel == 'IT':
        spa_list2 = ['Sorpresa', 'Noia', 'Interesse', 'Respirazione', 'Impulso']
    if lang_sel == 'EN':
        spa_list2 = ['Surprised', 'Boredom', 'Interest', 'Breath rate', 'Heart rate']
    if lang_sel == 'LT':
        spa_list2 = ['Nustebimas', 'Nuobodulys', 'Susidomėjimas', 'Kvėpavimo dažnis', 'Širdies ritmas']
    if lang_sel == 'GE':
        spa_list2 = ['Überrascht', 'Langeweile', 'Interesse', 'Atemfrequenz', 'Herzfrequenz']
    if lang_sel == 'FR':
        spa_list2 = ['Surpris', 'Ennui', 'Intérêt', 'Fréquence respiratoire', 'Fréquence cardiaque']
    pav_list2 = ['surprised', 'boredom', 'interest', 'breath', 'heart']
    df_normalized2 = normalizuoti_duomenis(df, pav_list2)
    vid_list2 = df_normalized2[pav_list2].mean().tolist()
    radar_fig2 = braizyti_radara(spa_list2, vid_list2)
    
    # Sukuriamas trecias grafikas
    if lang_sel == 'IT':
        spa_list3 = ['Aspettativa', 'Concentrazione', 'Pensiero positivo/Fiducia', 'Emotività', 'Vivacità', 'Passione', 'Premura']
    if lang_sel == 'EN':
        spa_list3 = ['Anticipation', 'Concentration', 'Confidence', 'Emotional', 'Energetic', 'Passionate', 'Thoughtful']
    if lang_sel == 'LT':
        spa_list3 = ['Laukimas', 'Susikaupimas', 'Pasitikėjimas savimi', 'Emocingumas', 'Energingumas', 'Aistringumas', 'Įžvalgumas']
    if lang_sel == 'GE':
        spa_list3 = ['Vorfreude', 'Konzentration', 'Selbstvertrauen', 'Emotional', 'Energisch', 'Leidenschaftlich', 'Nachdenklich']
    if lang_sel == 'FR':
        spa_list3 = ['Anticipation', 'Concentration', 'Confiance', 'Émotionnel', 'Énergique', 'Passionné', 'Réfléchi']
    pav_list3 = ['edp-anticipation', 'edp-concentrated', 'edp-confident', 'edp-emotional', 'edp-energetic', 'edp-passionate', 'edp-thoughtful']
    df_normalized3 = normalizuoti_duomenis(df, pav_list3)
    vid_list3 = df_normalized3[pav_list3].mean().tolist()
    radar_fig3 = braizyti_radara(spa_list3, vid_list3)
    
    # Sukuriamas ketvirtas grafikas
    if lang_sel == 'IT':
        spa_list4 = ['Stress', 'Coinvolgimento', 'Impulsività', 'Grinta', 'Agitazione', 'Immaginazione', 'Sforzo mentale']
    if lang_sel == 'EN':
        spa_list4 = ['Stress', 'Concentration', 'Impulsiveness', 'Energy', 'Excitement', 'Imagination', 'Mental Effort']
    if lang_sel == 'LT':
        spa_list4 = ['Stresas', 'Koncentracija', 'Impulsyvumas', 'Energija', 'Jaudulys', 'Vaizduotė', 'Protinės pastangos']
    if lang_sel == 'GE':
        spa_list4 = ['Stress', 'Konzentration', 'Impulsivität', 'Energie', 'Aufregung', 'Fantasie', 'Geistige Anstrengung']
    if lang_sel == 'FR':
        spa_list4 = ['Stress', 'Concentration', 'Impulsivité', 'Énergie', 'Excitation', 'Imagination', 'Effort mental']
    pav_list4 = ['emotionplayer-stress', 'concentration', 'emotioncognitiveratio', 'energy', 'excitement', 'imagination', 'mentaleffort']
    df_normalized4 = normalizuoti_duomenis(df, pav_list4)
    vid_list4 = df_normalized4[pav_list4].mean().tolist()
    radar_fig4 = braizyti_radara(spa_list4, vid_list4)
    
    # Sukuriamas penktas grafikas
    if lang_sel == 'IT':
        spa_list5 = ['Livello sonoro', 'Contenuto di CO2', 'Umidità ambientale', 'Illuminazione', "Pressione dell'aria", 'Contenuto di particolato']
    if lang_sel == 'EN':
        spa_list5 = ['Noise Level', 'CO2 concentration', 'Environmental Humidity', 'Lighting', 'Air pressure', 'Particulate Matter concentration']
    if lang_sel == 'LT':
        spa_list5 = ['Triukšmo lygis', 'CO2 koncentracija', 'Aplinkos drėgmė', 'Apšviestumas', 'Oro slėgis', 'Kietųjų dalelių kiekis']
    if lang_sel == 'GE':
        spa_list5 = ['Lärmpegel', 'CO2-Konzentration', 'Luftfeuchtigkeit', 'Beleuchtung', 'Luftdruck', 'Feinstaubkonzentration']
    if lang_sel == 'FR':
        spa_list5 = ['Niveau sonore', 'Concentration de CO2', 'Humidité ambiante', 'Éclairage', 'Press. atmosphérique', 'Concentration de particules']
    pav_list5 = ['ee_garso_lygis', 'steinel_co2', 'steinel_hih_2', 'steinel_lux_2', 'steinel_rel_air_pressure_2', 'steinel_voc_2']
    df_normalized5 = normalizuoti_duomenis(df, pav_list5)
    vid_list5 = df_normalized5[pav_list5].mean().tolist()
    radar_fig5 = braizyti_radara(spa_list5, vid_list5)
    
    # Konvertuojame visus grafikus i paveiksleliu duomenis atmintyje
    radar_image_data_list = []
    
    radar_image_data1 = BytesIO()
    radar_fig1.savefig(radar_image_data1, format='png', bbox_inches='tight')
    radar_image_data1.seek(0)
    radar_image_data_list.append(radar_image_data1)
    plt.close(radar_fig1)
 
    radar_image_data2 = BytesIO()
    radar_fig2.savefig(radar_image_data2, format='png', bbox_inches='tight')
    radar_image_data2.seek(0)
    radar_image_data_list.append(radar_image_data2)
    plt.close(radar_fig2)
 
    radar_image_data3 = BytesIO()
    radar_fig3.savefig(radar_image_data3, format='png', bbox_inches='tight')
    radar_image_data3.seek(0)
    radar_image_data_list.append(radar_image_data3)
    plt.close(radar_fig3)
 
    radar_image_data4 = BytesIO()
    radar_fig4.savefig(radar_image_data4, format='png', bbox_inches='tight')
    radar_image_data4.seek(0)
    radar_image_data_list.append(radar_image_data4)
    plt.close(radar_fig4)
 
    radar_image_data5 = BytesIO()
    radar_fig5.savefig(radar_image_data5, format='png', bbox_inches='tight')
    radar_image_data5.seek(0)
    radar_image_data_list.append(radar_image_data5)
    plt.close(radar_fig5)
 
    df_for_reg = df.copy()
    if 'time' in df_for_reg.columns:
        df_for_reg = df_for_reg.drop(columns=['time'])
    
    df_for_reg = df_for_reg.apply(pd.to_numeric, errors='coerce')
    df_for_reg = df_for_reg.dropna()
    
    all_columns = df_for_reg.columns.tolist()
    dependent_variables = ['angry', 'arousal', 'disgusted', 'neutral', 'sad_x', 'scared', 'surprised', 'valence', 'boredom', 'interest', 'breath', 'heart']
    dependent_variables = [var for var in dependent_variables if var in all_columns]
    
    if not dependent_variables:
        sys.exit(1)
 
    excel_file_path = os.path.splitext(csv_file_path)[0] + '.xlsx'
        
    wb = Workbook()
    
    default_sheet = wb.active
    if default_sheet.title == 'Sheet':
        wb.remove(default_sheet)
 
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold_font = Font(bold=True)
    normal_font = Font(bold=False)
    center_align = Alignment(horizontal="center")
    right_align = Alignment(horizontal="right")
    left_align = Alignment(horizontal="left")
 
    pdf_results = {}
    excel_plot_paths = []
 
    for dependent_var in dependent_variables:
        ws = wb.create_sheet(title=dependent_var)
        independent_vars = [col for col in all_columns if col != dependent_var]
        
        plot_path = create_and_save_time_plot(df_full, dependent_var, excel_file_path)
        if plot_path:
            excel_plot_paths.append(plot_path)
            img = OpenpyxlImage(plot_path)
            ws.add_image(img, 'B8')
 
        row_offset = 0
        
        while independent_vars:
            X = df_for_reg[independent_vars]
            X = sm.add_constant(X)
            y = df_for_reg[dependent_var]
            
            model = sm.OLS(y, X, missing='drop').fit()
            
            ws.row_dimensions[1 + row_offset].height = 15.0
            ws.row_dimensions[2 + row_offset].height = 15.0
            ws.row_dimensions[3 + row_offset].height = 15.0
            ws.row_dimensions[4 + row_offset].height = 15.0
            ws.row_dimensions[5 + row_offset].height = 15.0
            ws.row_dimensions[6 + row_offset].height = 15.0
 
            ws.cell(row=1 + row_offset, column=1, value='R2=').font = bold_font
            ws.cell(row=1 + row_offset, column=1).alignment = right_align
            ws.cell(row=1 + row_offset, column=2, value=model.rsquared).font = normal_font
            ws.cell(row=1 + row_offset, column=3, value='p=').font = bold_font
            ws.cell(row=1 + row_offset, column=3).alignment = right_align
            ws.cell(row=1 + row_offset, column=4, value=model.f_pvalue).font = normal_font
 
            ws.cell(row=2 + row_offset, column=1, value='Variables').font = bold_font
            ws.cell(row=2 + row_offset, column=1).alignment = left_align
            
            for col_idx, var_name in enumerate(model.params.index):
                cell = ws.cell(row=2 + row_offset, column=col_idx + 2, value=var_name)
                cell.font = bold_font
                cell.alignment = center_align
                if model.pvalues[var_name] < 0.05:
                    cell.fill = yellow_fill
 
            ws.cell(row=3 + row_offset, column=1, value='Coeficient B').font = bold_font
            ws.cell(row=3 + row_offset, column=1).alignment = left_align
            for col_idx, coef in enumerate(model.params):
                cell = ws.cell(row=3 + row_offset, column=col_idx + 2, value=coef)
                cell.font = normal_font
                cell.alignment = right_align
                if model.pvalues.iloc[col_idx] < 0.05:
                    cell.fill = yellow_fill
 
            ws.cell(row=4 + row_offset, column=1, value='p').font = bold_font
            ws.cell(row=4 + row_offset, column=1).alignment = left_align
            for col_idx, p_value in enumerate(model.pvalues):
                cell = ws.cell(row=4 + row_offset, column=col_idx + 2, value=p_value)
                cell.font = normal_font
                cell.alignment = right_align
                if p_value < 0.05:
                    cell.fill = yellow_fill
            
            ws.cell(row=5 + row_offset, column=1, value='Std. beta').font = bold_font
            ws.cell(row=5 + row_offset, column=1).alignment = left_align
            try:
                beta_standardized = (model.params.drop('const') * df_for_reg[independent_vars].std()) / df_for_reg[dependent_var].std()
                beta_standardized['const'] = None
                for col_idx, var_name in enumerate(model.params.index):
                    cell = ws.cell(row=5 + row_offset, column=col_idx + 2, value=beta_standardized.get(var_name))
                    cell.font = normal_font
                    cell.alignment = right_align
                    if model.pvalues[var_name] < 0.05:
                        cell.fill = yellow_fill
            except:
                pass
 
            ws.cell(row=6 + row_offset, column=1, value='Elasticity').font = bold_font
            ws.cell(row=6 + row_offset, column=1).alignment = left_align
            elasticities = calculate_elasticity(df_for_reg, dependent_var, independent_vars)
            
            for col_idx, var_name in enumerate(model.params.index):
                cell = ws.cell(row=6 + row_offset, column=col_idx + 2, value=elasticities.get(var_name))
                cell.font = normal_font
                cell.alignment = right_align
                if model.pvalues[var_name] < 0.05:
                    cell.fill = yellow_fill
            
            insignificant_vars = [var for var in model.pvalues.index if model.pvalues[var] > 0.05 and var != 'const']
 
            if not insignificant_vars:
                try:
                    beta_standardized_series = (model.params.drop('const') * df_for_reg[independent_vars].std()) / df_for_reg[dependent_var].std()
                except KeyError:
                    beta_standardized_series = (model.params * df_for_reg[independent_vars].std()) / df_for_reg[dependent_var].std()
                
                pdf_results[dependent_var] = {
                    'r2': model.rsquared,
                    'f_pvalue': model.f_pvalue,
                    'model_params': model.params,
                    'p_values': model.pvalues,
                    'elasticities': elasticities,
                    'beta_standardized': beta_standardized_series
                }
                break
            else:
                independent_vars = [var for var in independent_vars if var not in insignificant_vars]
                row_offset += 8
                if not independent_vars:
                    break
    
    try:
        wb.save(excel_file_path)
    except Exception as e:
        print(f"Klaida issaugant Excel faila: {e}")
        pass
    
    create_and_save_pdf(pdf_results, excel_file_path, df_full)
    
    for path in excel_plot_paths:
        try:
            os.remove(path)
        except OSError:
            pass
            
    try:
        print("\nPradedame el. laisko siuntimo procedura...")
        recipient_email = os.path.splitext(os.path.basename(csv_file_path))[0]
        if lang_sel == 'IT':
            subject = "Risultati esperimenti AFFECTS"
        if lang_sel == 'EN':
            subject = "Results of AFFECTS experiment"
        if lang_sel == 'LT':
            subject = "AFFECTS eksperimento rezultatai"
        if lang_sel == 'GE':
            subject = "Ergebnisse des AFFECTS-Experiments"
        if lang_sel == 'FR':
            subject = "Résultats de l'expérience AFFECTS"
        body = "Thank You for participation"
        
        excel_file_path = os.path.splitext(csv_file_path)[0] + '.xlsx'
        pdf_file_path = os.path.splitext(csv_file_path)[0] + '.pdf'
        zip_file_path = os.path.splitext(csv_file_path)[0] + '_ANOM.zip'
        
        attachment_paths = [excel_file_path, pdf_file_path, zip_file_path, excel_to_save_path]

        # Normalizuoja 'edp-confident' stulpelį
        min_val = df['edp-confident'].min()
        max_val = df['edp-confident'].max()

        # Apsauga nuo dalijimo iš nulio, jei visos stulpelio reikšmės vienodos
        if max_val - min_val > 0:
            df['EDP-Confident_normalized'] = (df['edp-confident'] - min_val) / (max_val - min_val)
        else:
            df['EDP-Confident_normalized'] = 0.5

        # Paskaičiuoja normalizuotų reikšmių vidurkį
        conf_val = df['EDP-Confident_normalized'].mean()

        # Priskiria reikšmę kintamajam "confidence"
        if conf_val >= 0.5:
            if lang_sel == 'IT':
                confidence = 'Nel corso della visione di questo video, il sistema ha rilevato che sei una persona che tende al pensiero positivo.'
            if lang_sel == 'EN':
                confidence = 'While watching this video, the system detected that you are a person who tends to think positively.'
            if lang_sel == 'LT':
                confidence = 'Žiūrint šį vaizdo įrašą, sistema aptiko, kad esate linkęs mąstyti pozityviai.'
            if lang_sel == 'GE':
                confidence = 'Während Sie dieses Video ansehen, hat das System festgestellt, dass Sie zu einer positiven Denkweise neigen.'
            if lang_sel == 'FR':
                confidence = 'En visionnant cette vidéo, le système a détecté que vous êtes une personne qui a tendance à penser positivement.'
        else:
            if lang_sel == 'IT':
                confidence = 'Nel corso della visione di questo video, il sistema ha rilevato che questo non ha generato in te un pensiero positivo.'
            if lang_sel == 'EN':
                confidence = 'While watching this video, the system detected that it did not generate a positive thought in you.'
            if lang_sel == 'LT':
                confidence = 'Žiūrint šį vaizdo įrašą, sistema aptiko, kad jis jumyse nesukėlė teigiamų emocijų.'
            if lang_sel == 'GE':
                confidence = 'Während Sie dieses Video ansahen, stellte das System fest, dass es bei Ihnen keine positiven Gedanken auslöste.'
            if lang_sel == 'FR':
                confidence = "Lors du visionnage de cette vidéo, le système a détecté qu'elle ne générait pas de pensée positive en vous."

        send_email_with_attachments(recipient_email, subject, body, attachment_paths, radar_image_data_list, confidence, lang_sel)
 
    except Exception as e:
        print(f"Klaida ruosiant el. laisko siuntima: {e}")
 
if __name__ == "__main__":
    main()